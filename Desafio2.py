import openpyxl
import os
from openpyxl.utils import get_column_letter

def update_excel(file_path):
    

    if not file_path.endswith('.xlsx'):
        print("Error: El archivo debe tener la extensión '.xlsx'.")
        return
    
    
    if not os.path.exists(file_path):
        print("El archivo no existe. Creando un nuevo archivo...")
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Nombre", "Nota", "Asignatura"])  
        wb.save(file_path)
    else:
        print("El archivo ya existe. Agregando nueva fila de datos...")

    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active


    new_row = ["Juan", "80", "Matemáticas"]
    sheet.append(new_row)

    
    for col in sheet.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    
    try:
        wb.save(file_path)
        print("Los datos fueron agregados correctamente al archivo Excel.")
    except PermissionError:
        print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")
    
    
    wb.close()


file_path = "notas.xlsx"
update_excel(file_path)
